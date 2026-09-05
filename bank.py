"""本地 Excel 题库的载入与记录"""
import openpyxl

from utils import text_format


class QuestionBank:
    def __init__(self, subject_id):  # 载入文件(不存在则进行初始化)
        self.path = f'青马易战_{subject_id}.xlsx'
        try:
            self.workbook = openpyxl.load_workbook(self.path)
            self.sheet = self.workbook.active
        except FileNotFoundError:
            self.workbook = openpyxl.Workbook()
            self.sheet = self.workbook.active
            data_list = ['题目', '选项', '正确选项', '正确答案']
            self.sheet.append(data_list)
        self.sheet.column_dimensions['A'].width = 60
        self.sheet.column_dimensions['B'].width = 80
        self.sheet.column_dimensions['C'].width = 10
        self.sheet.column_dimensions['D'].width = 80
        # 加载xlsx题库到变量questions
        self.questions = {}
        for row in self.sheet.iter_rows(values_only=True):
            self.questions[text_format(row[0])] = row[2]

    def record(self, question, text_options, answer, right_answer):  # 若题目不在本地题库中则加入本地题库
        if question in self.questions:
            return
        self.questions[question] = answer
        data_list = [question, text_options, answer, right_answer]
        self.sheet.append(data_list)
        self.workbook.save(self.path)
