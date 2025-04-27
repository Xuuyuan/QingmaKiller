import requests
import time
import json
from Crypto.Cipher import AES
from Crypto.Util.Padding import unpad
import openpyxl
import random
import base64
import re
from bs4 import BeautifulSoup

header_login = {
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
    'Accept-Encoding': 'gzip, deflate',
    'Accept-Language': 'zh-CN,zh;q=0.9',
    'Host': '112.5.88.114:31101',
    'Proxy-Connection': 'keep-alive',
    'Upgrade-Insecure-Requests': '1',
    'User-Agent': 'Mozilla/5.0 (Linux; Android 10; HLK-AL00 Build/HONORHLK-AL00; wv) AppleWebKit/537.36 (KHTML, like Gecko) Version/4.0 Chrome/78.0.3904.108 Mobile Safari/537.36 yiban_android',
}


def text_format(text):  # 清除不必要的字符 将unicode文本格式化为字符串以便于搜题
    text = str(text.replace('\u3000', '').replace('\xa0', ''))
    for i in [' ', '“', '”', '"', '.', '．']:
        text = text.replace(i, '')
    return text


def gettime():  # 获取现行时间戳
    t = str(time.time())
    return t[0:10] + t[11:14]


def decrypt(text):  # 解密青马易战文本 返回值类型为unicode文本
    if text == '':
        return ''
    key = base64.b64decode("ZDBmMTNiZGI3MDRhMWVhMWE3MTcwNjJiNTk0NzY0ODg=")
    cipher = AES.new(key, AES.MODE_ECB)
    decrypted = cipher.decrypt(base64.b64decode(text))
    decrypted = unpad(decrypted, AES.block_size, style='pkcs7').decode('utf-8')
    return decrypted


def get_course_list(cookie) -> dict:  # 获取课程列表
    headers = {
        'Referer': f'http://112.5.88.114:31101/yiban-web/stu/toCourse.jhtml',
        'User-Agent': 'Mozilla/5.0 (Linux; Android 10; HLK-AL00 Build/HONORHLK-AL00; wv) AppleWebKit/537.36 (KHTML, like Gecko) Version/4.0 Chrome/78.0.3904.108 Mobile Safari/537.36 yiban_android',
        'Cookie': cookie,
        'Accept': 'application/json',
        'Origin': 'http://112.5.88.114:31101',
        'Host': '112.5.88.114:31101'
    }

    response = requests.get("http://112.5.88.114:31101/yiban-web/stu/toCourse.jhtml",
                            data="", headers=headers, allow_redirects=False)
    if response.status_code == 302:
        # print(response.headers['Location'])
        return {"isSuccess": False}
    soup = BeautifulSoup(response.text, 'html.parser')
    courses = {}
    for li in soup.find_all('li', class_='mui-table-view-cell mui-media mui-col-xs-6 mui-col-sm-6 course-li'):
        a_tag = li.find('a', class_='ahref')
        course_id = re.search(r'courseId=(\d+)', a_tag['href']).group(1)
        course_name = li.find('div', class_='mui-media-body').text.strip()
        courses[course_id] = course_name
    return {"isSuccess": True, "courses": courses}


def get_cookie_from_url(url):
    res = requests.get(url, headers=header_login)
    cookie = re.match(r'JSESSIONID=\w*', res.headers['set-cookie']).group()
    return cookie


options_list = ['A', 'B', 'C', 'D', 'E', 'F', 'G']
options_dict = {'A': 0, 'B': 1, 'C': 2, 'D': 3, 'E': 4, 'F': 5, 'G': 6}
headers_tiku = {"content-type": "application/json"}


def main():
    print('====Qingmakiller 青马易战自动答题工具====')
    # 获取cookie
    url = input('请输入青马易战URL: ')
    for _ in range(5):
        cookie = get_cookie_from_url(url)
        if cookie:
            print('# 获取到cookie: ' + cookie)
        else:
            print('# 获取cookie失败! ')
            return

        # 获取课程列表
        course_list = get_course_list(cookie)
        if course_list['isSuccess']:  # 获取成功
            break
        else:
            print('# 被跳转到授权页面，将在0.5秒后自动重试..')
            time.sleep(0.5)
            continue
    if not course_list['isSuccess']:
        print('# 被跳转到授权页面，请检查URL是否过期! （经测试，正常的链接也有概率跳转到授权页面，可以尝试重复此操作。）')
        return

    # 科目列表展示
    print("科目ID | 科目名称")
    for i in course_list['courses'].items():
        print(f"{i[0]} | {i[1]}")

    # 选定需要刷题的科目
    subjectId = input('请输入需要刷题的科目ID: ')
    if subjectId == '':
        print('# 未输入科目ID! ')
        return
    now_times_str = input('请输入当前答题数（不输入默认为0）: ')
    now_right_times_str = input('请输入当前答对数（不输入默认为0）: ')
    now_times = int(now_times_str) if now_times_str != '' else 0
    now_right_times = int(
        now_right_times_str) if now_right_times_str != '' else 0
    now_right_rate = now_right_times / now_times if now_times != 0 else 0  # 计算当前正确率

    target_times_str = input(
        '请输入需要刷到的保底答对次数（保底答对次数=本次运行答对数-当前答对数, 不输入默认为550）: ')
    target_times = int(target_times_str) if target_times_str != '' else 550

    target_right_rate_str = input('请输入需要刷题的保底正确率（不输入默认为0.6）: ')
    target_right_rate = float(
        target_right_rate_str) if target_right_rate_str != '' else 0.6

    max_right_rate_str = input('请输入需要刷题的上限正确率（不输入默认为0.9）: ')
    max_right_rate = float(
        max_right_rate_str) if max_right_rate_str != '' else 0.9

    headers = {
        'Referer': f'http://112.5.88.114:31101/yiban-web/stu/toSubject.jhtml?courseId={subjectId}',
        'User-Agent': 'Mozilla/5.0 (Linux; Android 10; HLK-AL00 Build/HONORHLK-AL00; wv) AppleWebKit/537.36 (KHTML, like Gecko) Version/4.0 Chrome/78.0.3904.108 Mobile Safari/537.36 yiban_android',
        'Cookie': cookie,
        'Accept': 'application/json',
        'Origin': 'http://112.5.88.114:31101',
        'Host': '112.5.88.114:31101'
    }

    # 载入文件（不存在则进行初始化）
    try:
        workbook = openpyxl.load_workbook(f'青马易战_{subjectId}.xlsx')
        sheet = workbook.active
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        data_list = ['题目', '选项', '正确选项', '正确答案']
        sheet.append(data_list)
    sheet.column_dimensions['A'].width = 60
    sheet.column_dimensions['B'].width = 80
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 80
    # 加载xlsx题库到变量questions
    questions = {}
    for row in sheet.iter_rows(values_only=True):
        questions[text_format(row[0])] = row[2]

    run_times = 0
    # 开始运行
    while now_right_times <= target_times or now_right_rate < target_right_rate:  # 循环条件
        # 获取题目及选项
        req = requests.post(
            f'http://112.5.88.114:31101/yiban-web/stu/nextSubject.jhtml?_={gettime()}', headers=headers, data={'courseId': subjectId})
        # TODO
        if 'document.location=\'/host_not_found_error\'' in req.text:
            print('# 该URL已过期, 请根据指引重新获取URL! ')
            break
        html = json.loads(req.text)
        if 'uuid' not in html['data']:
            print('# 题目获取失败! 返回信息: ', html)
            break
        UUID = html['data']['uuid']
        now_question = text_format(
            decrypt(html['data']['nextSubject']['subDescript']))
        now_question_type = 1 if html['data']['nextSubject']['subType'] == '多选题' else 0
        # 0为单选 1为多选
        options = [text_format(decrypt(html['data']['nextSubject'][f'option{i}'])) for i in range(
            html['data']['nextSubject']['optionCount']) if f'option{i}' in html['data']['nextSubject']]
        text_options = ' '.join(
            options_list[i] + '. ' + options[i] + ' ' for i in range(len(options)))
        print(f'{now_right_times}->{target_times}',
              decrypt(html['data']['nextSubject']['subDescript']), text_options)
        # 检测刷题题目
        if '刷题' in now_question or '请选择' in now_question:
            print('# 检测到防刷题题目, 自动跳过')
            continue
        if now_question in questions:  # 题目已存在本地题库
            if now_right_rate > max_right_rate:  # 正确率过高
                print('# 该题目已存在本地题库中/正确率过高!  正确答案为 ' +
                      questions[now_question] + ', 将自动提交随机答案! ')
                if now_question_type == 0:
                    my_answer = options_list[random.randint(0, len(options)-1)]
                else:
                    my_answer = ''.join(random.sample(
                        options_list[:len(options)], 2))
            else:
                print('# 该题目已存在本地题库中!  正确答案为 ' +
                      questions[now_question] + ', 将自动提交! ')
                my_answer = questions[now_question]
        else:  # 题目不存在本地题库中, 尝试从网络题库获取
            # my_answer = ''
            force_choose = False
            if now_question_type == 0:
                for d in range(len(options)):
                    if '下都是' in options[d] or '上都是' in options[d]:  # 有出现对应文本的选项直接提交
                        my_answer = options_list[d]
                        force_choose = True
            if now_right_rate > max_right_rate:  # 正确率过高, 随机提交答案
                if now_question_type == 0:  # 单选题
                    my_answer = options_list[random.randint(0, len(options)-1)]
                else:  # 多选题
                    my_answer = ''.join(random.sample(
                        options_list[:len(options)], 2))
                print('# 正确率过高, 将自动提交随机答案! ')
                force_choose = True
            elif not force_choose:
                try:
                    req_tiku = requests.post('http://localhost:8060/adapter-service/search', json={
                                             "question": now_question, "type": now_question_type, "options": options}, headers=headers_tiku)
                except requests.exceptions.ConnectionError:
                    print('# 无法连接到tikuAdapter, 自动跳过本题, 请先启动tikuAdapter再运行本程序! ')
                    time.sleep(random.randint(4, 10))
                    continue
                except Exception as e:
                    print('# 连接tikuAdapter时发现错误, 自动跳过本题: ', e)
                    time.sleep(random.randint(4, 10))
                    continue
                res_tiku = json.loads(req_tiku.text)
                my_answer = res_tiku['answer']['answerKeyText']
                if now_question_type == 0 and len(my_answer) > 1:  # 避免单选题提交多选
                    my_answer = my_answer[0]
                # 避免多选题提交单选
                elif now_question_type == 1 and len(my_answer) == 1:
                    my_answer = ''

                if now_question_type == 1:  # 多选题去重
                    my_answer = ''.join(set(my_answer))
                print('# 网络题库搜索结果: ' + my_answer)
            if my_answer == '':  # 网络题库未找到答案
                if now_right_rate > max_right_rate:  # 正确率高于保底正确率, 随机提交
                    if now_question_type == 0:
                        my_answer = options_list[random.randint(
                            0, len(options)-1)]
                    else:
                        my_answer = ''.join(random.sample(
                            options_list[:len(options)], 2))
                    time.sleep(1)
                else:
                    print('# 网络题库未找到答案, 本题跳过! ')
                    time.sleep(3)
                    continue
        # 随机休眠, 防止检测
        time.sleep(random.randint(4, 10))

        # 提交答案
        data_submit = {'answer': my_answer,
                       'courseId': subjectId, 'uuid': UUID, 'deviceUuid': ""}
        req_submit = requests.post(
            f'http://112.5.88.114:31101/yiban-web/stu/changeSituation.jhtml?_={gettime()}', headers=headers, data=data_submit)
        res_submit = json.loads(req_submit.text)

        if res_submit['message'] == '回答正确！':  # 回答正确
            run_times += 1
            now_times += 1
            now_right_times += 1
            now_right_rate = now_right_times / now_times
            if now_question not in questions:  # 若题目不在本地题库中则加入本地题库
                data_list = [now_question, text_options, my_answer, '?']
                sheet.append(data_list)
                workbook.save(f'青马易战_{subjectId}.xlsx')
                questions[now_question] = my_answer
            print(
                f'# 本题回答正确!  当前提交次数 {run_times} 目标答对数 {target_times} 现答对数 {now_right_times} 现答题数 {now_times} 正确率 {now_right_rate * 100:.2f}%/{target_right_rate * 100}%')
        elif res_submit['message'] == '回答错误！':  # 回答错误
            run_times += 1
            now_times += 1
            now_right_rate = now_right_times / now_times
            rightAnswer = decrypt(res_submit['data']['rightOption'])
            if now_question not in questions:  # 若题目不在本地题库中则加入本地题库
                questions[now_question] = ''
                for k in options_list:
                    if k in rightAnswer:
                        questions[now_question] += k
                data_list = [now_question, text_options,
                             questions[now_question], rightAnswer]
                sheet.append(data_list)
                workbook.save(f'青马易战_{subjectId}.xlsx')

            print(f'# 本题回答错误!  提交答案: {my_answer} 正确答案: {rightAnswer} 当前提交次数 {run_times}/{target_times} 现答对数 {now_right_times} 现答题数 {now_times} 正确率 {now_right_rate * 100:.2f}%/{target_right_rate * 100}%')
        elif res_submit['message'] == '您的答题速度过快，请认真答题，30s后可继续答题.':  # 触发答题冷却
            print('# 触发答题冷却, 等待30s后自动继续..')
            time.sleep(30)
        else:  # 发现错误, 等待手动处理
            print(f'# 发现未知回复文本: {res_submit}')
            input()
    print('# 运行完毕。')


if __name__ == '__main__':
    while True:
        try:
            main()
        except Exception as e:
            print('# 发现错误: ', e)
